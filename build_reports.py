"""
build_reports.py (v2)
─────────────────────
FIXY vs v1:
  1. Czyste nazwy klubów/teamów z teamy_kluby_25_26.csv
     (final_club_name, final_team_name)
  2. Klub w aktualnym sezonie = chronologicznie ostatni (last_club_id z SQL).
     Jeśli zmiana klubu w sezonie → "Klub A → Klub B"
  3. Sezon debiutu wybierany przez debut_season_id (z pro_paths v3),
     nie przez years_before_debut==0 (bo to było buggy)
  4. Tylko pro-gracze z ≥3 sezonami w arkuszu "Pro debiut"
  5. Trajectory matching: WIEK-DO-WIEKU (kandydat 18-19-20 vs pro 18-19-20)
     zamiast sezon_kandydata vs sezon_pro_przed_debiutem

Wejście (data/):
  pro_paths.csv (v3), candidates.csv (v4), top_score.csv,
  teamy_kluby_25_26_-_teamy_kluby_25_26.csv

Wyjście (data/):
  pro_career_paths.xlsx, candidate_matches.xlsx
"""

from pathlib import Path
import time
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import LEVEL_NAMES

DATA_DIR = Path("data")

# ══════════════════════════════════════════════════════════════════════════════
# STYLE
# ══════════════════════════════════════════════════════════════════════════════

FONT_HEADER  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_BODY    = Font(name="Arial", size=10)
FONT_BOLD    = Font(name="Arial", size=10, bold=True)
FONT_TITLE   = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_MATCH_HIGHLIGHT = Font(name="Arial", size=10, bold=True, color="C00000")
FONT_SECTION = Font(name="Arial", size=11, bold=True)

FILL_HEADER = PatternFill("solid", start_color="2F5597")
FILL_CAND   = PatternFill("solid", start_color="FFF2CC")
FILL_MAIN   = PatternFill("solid", start_color="D9E1F2")
FILL_BACKUP = PatternFill("solid", start_color="EDEDED")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

MIN_SEASONS_FOR_PRO = 3  # filtr w arkuszu "Pro debiut"


def lvl_name(n):
    if pd.isna(n):
        return "—"
    return LEVEL_NAMES.get(int(n), f"poziom {int(n)}")


def activity_from_pct(pct):
    if pd.isna(pct):  return "nieznana"
    if pct >= 70:     return "podstawowy"
    if pct >= 40:     return "regularny"
    if pct >= 15:     return "rezerwowy"
    return "sporadyczny"


# ══════════════════════════════════════════════════════════════════════════════
# 1. WCZYTAJ DANE + MAPOWANIE CZYSTYCH NAZW
# ══════════════════════════════════════════════════════════════════════════════

def load_csv(path):
    if not Path(path).exists():
        raise FileNotFoundError(f"Brak pliku: {path}")
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
    df.columns = (
        df.columns.str.strip()
                  .str.replace("\ufeff", "", regex=False)
                  .str.lower()
                  .str.replace(" ", "_")
    )
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].str.strip()
        try:
            df[col] = df[col].str.replace(",", ".").astype(float)
        except Exception:
            pass
    return df


print("Wczytuję dane...")
pro_paths  = load_csv(DATA_DIR / "pro_paths.csv")
candidates = load_csv(DATA_DIR / "candidates.csv")
top_score  = load_csv(DATA_DIR / "top_score.csv")

# Mapowanie czystych nazw
NAME_MAP_PATH = DATA_DIR / "teamy_kluby_25_26_-_teamy_kluby_25_26.csv"
if NAME_MAP_PATH.exists():
    name_map = load_csv(NAME_MAP_PATH)
    # club_id → final_club_name (deduplikuj — bierz pierwszy non-null)
    club_map = (
        name_map.dropna(subset=["club_id", "final_club_name"])
        .drop_duplicates("club_id")
        .set_index("club_id")["final_club_name"]
        .to_dict()
    )
    team_map = (
        name_map.dropna(subset=["team_id", "final_team_name"])
        .drop_duplicates("team_id")
        .set_index("team_id")["final_team_name"]
        .to_dict()
    )
    print(f"  Mapowanie: {len(club_map)} klubów, {len(team_map)} teamów")
else:
    club_map = {}
    team_map = {}
    print("  UWAGA: brak teamy_kluby_25_26.csv — używam nazw z bazy")


def clean_club(cid, fallback):
    """Spróbuj clean name, fallback do oryginalnej."""
    if pd.notna(cid) and cid in club_map:
        return club_map[cid]
    return fallback if pd.notna(fallback) else "—"


def clean_team(tid, fallback):
    if pd.notna(tid) and tid in team_map:
        return team_map[tid]
    return fallback if pd.notna(fallback) else "—"


# Normalizacja boolean
for df in [pro_paths, candidates]:
    if "is_junior_league" in df.columns:
        df["is_junior_league"] = (
            df["is_junior_league"].astype(str).str.lower().isin(["true", "1", "1.0"])
        )

# ── Zastosuj czyste nazwy ─────────────────────────────────────────────────────
# pro_paths ma już club_name/team_name z fallback last→dominant (z SQL v3)
# Jeśli mamy club_id, mapujemy. Jeśli nie — zostaje co jest.
if "club_id" in pro_paths.columns:
    pro_paths["club_name"] = pro_paths.apply(
        lambda r: clean_club(r.get("club_id"), r.get("club_name")), axis=1
    )
if "team_id" in pro_paths.columns:
    pro_paths["team_name"] = pro_paths.apply(
        lambda r: clean_team(r.get("team_id"), r.get("team_name")), axis=1
    )

# Candidates ma TWA pary: last_* (preferred) i dominant_* (fallback)
# Wybieramy: dla każdej kolumny last → jeśli NaN → dominant
def pick_club_team(df):
    if "last_club_id" in df.columns and "dominant_club_id" in df.columns:
        df["club_id"] = df["last_club_id"].fillna(df["dominant_club_id"])
        df["club_name_raw"] = df["club_name_last"].fillna(df["club_name_dominant"])
        df["team_id"] = df["last_team_id"].fillna(df["dominant_team_id"])
        df["team_name_raw"] = df["team_name_last"].fillna(df["team_name_dominant"])
        # Także DOMINANT — żeby pokazać "X → Y" jeśli różne
        df["club_name_dominant_clean"] = df.apply(
            lambda r: clean_club(r.get("dominant_club_id"), r.get("club_name_dominant")), axis=1
        )
        df["club_name_last_clean"] = df.apply(
            lambda r: clean_club(r.get("last_club_id"), r.get("club_name_last")), axis=1
        )
        df["club_name"] = df.apply(
            lambda r: clean_club(r.get("club_id"), r.get("club_name_raw")), axis=1
        )
        df["team_name"] = df.apply(
            lambda r: clean_team(r.get("team_id"), r.get("team_name_raw")), axis=1
        )
    elif "club_id" in df.columns:
        df["club_name"] = df.apply(
            lambda r: clean_club(r.get("club_id"), r.get("club_name")), axis=1
        )
        if "team_id" in df.columns:
            df["team_name"] = df.apply(
                lambda r: clean_team(r.get("team_id"), r.get("team_name")), axis=1
            )
    return df

candidates = pick_club_team(candidates)

print(f"  Pro paths:  {len(pro_paths)} wierszy, {pro_paths['player_id'].nunique()} graczy")
print(f"  Candidates: {len(candidates)} wierszy, {candidates['player_id'].nunique()} graczy")
print(f"  Top score:  {len(top_score)} graczy referencyjnych")


# ══════════════════════════════════════════════════════════════════════════════
# 2. AKTYWNOŚĆ — wspólne max minut per liga × sezon
# ══════════════════════════════════════════════════════════════════════════════

print("\nLiczę aktywność...")

combined_min = pd.concat([
    pro_paths[["league_id", "season_id", "player_id", "total_minutes"]],
    candidates[["league_id", "season_id", "player_id", "total_minutes"]],
]).drop_duplicates(["league_id", "season_id", "player_id"])

max_per = (combined_min.groupby(["league_id", "season_id"])["total_minutes"]
           .quantile(0.95).reset_index(name="max_minutes_local"))

for df_name, df in [("candidates", candidates), ("pro_paths", pro_paths)]:
    # Usuń ewentualne stare kolumny (z poprzednich uruchomień / SQL)
    df.drop(columns=["max_minutes_local", "pct_max_minutes", "activity_status"],
            errors="ignore", inplace=True)
    merged = df.merge(max_per, on=["league_id", "season_id"], how="left")
    df["pct_max_minutes"] = (
        merged["total_minutes"] / merged["max_minutes_local"] * 100
    ).round(1).clip(0, 100)
    df["activity_status"] = df["pct_max_minutes"].apply(activity_from_pct)
    if df_name == "candidates":
        candidates = df
    else:
        pro_paths = df

print(f"  Kandydaci activity: {candidates['activity_status'].value_counts().to_dict()}")


# ══════════════════════════════════════════════════════════════════════════════
# 3. AGREGACJA PER SEZON
# ══════════════════════════════════════════════════════════════════════════════

def aggregate_seasons(df, is_pro=False):
    """
    Jeden wiersz na (player, season). highest_level = MAX poziomu w sezonie.
    club_name → klub z najwyższego poziomu (preferowane) lub z największą
    liczbą minut (fallback). Jeśli był zmiany klubu w sezonie i top-poziom-klub
    różny od dominant-minut-klub, pokażemy "A → B".
    """
    df = df.copy()
    if "is_junior_league" not in df.columns:
        df["is_junior_league"] = False

    df_sorted = df.sort_values(
        ["player_id", "season_id", "league_level", "total_minutes"],
        ascending=[True, True, False, False],  # NAJPIERW po lidze, potem minutach
    ).reset_index(drop=True)
    grp = ["player_id", "season_id"]

    df_sorted["league_blurb"] = (
        df_sorted["league_name"].astype(str)
        + " (" + df_sorted["total_minutes"].astype(int).astype(str)
        + " min, " + df_sorted["matches_played"].astype(int).astype(str) + "m)"
    )
    leagues_str = (
        df_sorted.groupby(grp)["league_blurb"]
                 .apply(" + ".join).reset_index(name="leagues_str")
    )

    # 'first' bierze wiersz z najwyższego poziomu (po sortowaniu)
    agg = df_sorted.groupby(grp, as_index=False).agg(
        age_in_season=("age_in_season", "max"),
        highest_level=("league_level", "max"),
        total_minutes=("total_minutes", "sum"),
        matches_played=("matches_played", "sum"),
        avg_score=("avg_score", "mean"),
        highest_is_junior=("is_junior_league", "first"),  # z najwyższego
        club_name=("club_name", "first"),                  # z najwyższego
        team_name=("team_name", "first"),
        activity_status=("activity_status", "first"),
        pct_max_minutes=("pct_max_minutes", "max"),
    )

    # Drugi klub w sezonie (jeśli inny) — żeby pokazać "A → B"
    # Dla kandydatów: porównaj z "klub z największą liczbą minut" (dominant)
    if "club_name_dominant_clean" in df_sorted.columns:
        # Klub dominujący per sezon
        dom = (df_sorted.sort_values(["player_id", "season_id", "total_minutes"],
                                     ascending=[True, True, False])
               .drop_duplicates(grp)
               [grp + ["club_name_dominant_clean"]]
               .rename(columns={"club_name_dominant_clean": "club_name_dominant"}))
        agg = agg.merge(dom, on=grp, how="left")
    else:
        agg["club_name_dominant"] = agg["club_name"]

    out = agg.merge(leagues_str, on=grp, how="left")

    if is_pro and "years_before_debut" in df.columns:
        ydb = df.groupby(grp, as_index=False).agg(
            years_before_debut=("years_before_debut", "first"),
            age_at_debut=("age_at_debut", "first"),
            player_name=("player_name", "first"),
            debut_season_id=("debut_season_id", "first"),
            debut_league_level=("debut_league_level", "first"),
        )
        out = out.merge(ydb, on=grp, how="left")
    elif "player_name" in df.columns:
        names = df.groupby("player_id", as_index=False)["player_name"].first()
        out = out.merge(names, on="player_id", how="left")
    return out


print("\nAgregacja sezonów pro...")
pro_seasons = aggregate_seasons(pro_paths, is_pro=True)
print(f"  {len(pro_seasons)} sezonów, {pro_seasons['player_id'].nunique()} pro-graczy")

print("Agregacja sezonów kandydatów...")
cand_seasons = aggregate_seasons(candidates, is_pro=False)
print(f"  {len(cand_seasons)} sezonów, {cand_seasons['player_id'].nunique()} kandydatów")

pro_by_player = {pid: g.sort_values("years_before_debut", ascending=False)
                 for pid, g in pro_seasons.groupby("player_id")}


# ══════════════════════════════════════════════════════════════════════════════
# 4. PRO_CAREER_PATHS.XLSX — z filtrem ≥3 sezonów
# ══════════════════════════════════════════════════════════════════════════════

print(f"\nBuduję pro_career_paths.xlsx (filtr: ≥{MIN_SEASONS_FOR_PRO} sezony)...")
t0 = time.time()


def build_pro_debut_sheet(wb):
    ws = wb.active
    ws.title = "Pro debiut"

    season_offsets = [0, -1, -2, -3, -4, -5]
    headers = ["Imię i nazwisko", "Wiek debiutu", "Klub debiutu", "Liga debiutu",
               "Liczba sezonów"]
    for offset in season_offsets:
        lbl = "Debiut" if offset == 0 else f"Sezon {offset}"
        headers += [f"{lbl}: Liga", f"{lbl}: Klub", f"{lbl}: Min", f"{lbl}: Mecze",
                    f"{lbl}: Aktywność"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_CENTER

    row_idx = 2
    n_written = 0
    for pid, grp in pro_by_player.items():
        # Filtr: minimum N sezonów
        if len(grp) < MIN_SEASONS_FOR_PRO:
            continue

        # Sezon debiutu = ten gdzie season_id == debut_season_id
        debut_season_id = grp["debut_season_id"].iloc[0]
        debut_row = grp[grp["season_id"] == debut_season_id]
        if len(debut_row) == 0:
            continue
        d = debut_row.iloc[0]
        debut_level = int(d["debut_league_level"])
        if debut_level not in (11, 12, 13):
            continue  # nie powinno się zdarzyć, ale safety

        cells = [
            d["player_name"],
            int(d["age_at_debut"]),
            str(d["club_name"]),
            lvl_name(debut_level),
            int(len(grp)),
        ]
        for offset in season_offsets:
            ybd = -offset  # 0=debiut, 1=rok przed, ...
            srow = grp[grp["years_before_debut"] == ybd]
            if len(srow) > 0:
                r = srow.iloc[0]
                cells += [
                    lvl_name(r["highest_level"]),
                    str(r["club_name"]) if pd.notna(r["club_name"]) else "—",
                    int(r["total_minutes"]),
                    int(r["matches_played"]),
                    str(r["activity_status"]),
                ]
            else:
                cells += ["—", "—", "", "", "—"]

        for col_idx, val in enumerate(cells, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT if isinstance(val, str) else ALIGN_CENTER
        row_idx += 1
        n_written += 1

    widths = [24, 9, 28, 22, 10]
    for _ in season_offsets:
        widths += [22, 24, 8, 8, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "F2"
    print(f"  Pro debiut: zapisano {n_written} pro-graczy (z ≥{MIN_SEASONS_FOR_PRO} sezonami)")


def build_pro_details_sheet(wb):
    ws = wb.create_sheet("Detale")
    headers = ["Player ID", "Imię i nazwisko", "Wiek debiutu",
               "Lat przed debiutem", "Wiek w sezonie",
               "Liga(i)", "Łącznie min", "Mecze",
               "Aktywność", "Klub", "Match score"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_CENTER

    row_idx = 2
    for pid, grp in pro_by_player.items():
        for _, r in grp.iterrows():
            score = r.get("avg_score")
            ws.append([
                str(pid),
                str(r["player_name"]),
                int(r["age_at_debut"]),
                int(r["years_before_debut"]),
                int(r["age_in_season"]),
                str(r["leagues_str"]),
                int(r["total_minutes"]),
                int(r["matches_played"]),
                str(r["activity_status"]),
                str(r["club_name"]) if pd.notna(r["club_name"]) else "—",
                round(float(score), 3) if pd.notna(score) else None,
            ])
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).font = FONT_BODY
            row_idx += 1

    for i, w in enumerate([38, 24, 9, 12, 9, 60, 12, 8, 14, 28, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"


def build_legend_sheet(wb):
    ws = wb.create_sheet("Legenda")
    ws["A1"] = "POZIOMY LIG (skala 1–13)"
    ws["A1"].font = FONT_SECTION
    row = 2
    for lvl in sorted(LEVEL_NAMES.keys(), reverse=True):
        ws.cell(row=row, column=1, value=lvl).font = FONT_BODY
        ws.cell(row=row, column=2, value=LEVEL_NAMES[lvl]).font = FONT_BODY
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="AKTYWNOŚĆ").font = FONT_SECTION
    for k, v in [("podstawowy", "≥ 70% maksymalnych minut sezonu w lidze"),
                 ("regularny", "40–70%"), ("rezerwowy", "15–40%"),
                 ("sporadyczny", "< 15%")]:
        row += 1
        ws.cell(row=row, column=1, value=k).font = FONT_BODY
        ws.cell(row=row, column=2, value=v).font = FONT_BODY
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60


wb_pro = Workbook()
build_pro_debut_sheet(wb_pro)
build_pro_details_sheet(wb_pro)
build_legend_sheet(wb_pro)
wb_pro.save(DATA_DIR / "pro_career_paths.xlsx")
print(f"  Zapisano: data/pro_career_paths.xlsx ({time.time()-t0:.1f}s)")


# ══════════════════════════════════════════════════════════════════════════════
# 5. TRAJECTORY MATCHING — WIEK-DO-WIEKU
# ══════════════════════════════════════════════════════════════════════════════
# Zmiana koncepcji: porównujemy "kandydat w wieku 18-19-20" z "pro w wieku 18-19-20".
# Czyli klucz jest WIEK, nie odległość od debiutu.
#
# Dla każdego pro budujemy mapę: age → (level, minutes, score).
# Dla kandydata bierzemy 3 ostatnie wieki. Szukamy pro który w TYCH SAMYCH wiekach
# miał najbliższy wektor.

print("\nLiczę trajectory matching (wiek-do-wieku)...")

top_ids = set(top_score["player_id"].tolist())

# pro_age_traj[pid][age] = (level/13, min/2000, score/0.7)
pro_age_traj = {}
for pid, grp in pro_seasons.groupby("player_id"):
    if pid not in top_ids:
        continue
    traj = {}
    for _, r in grp.iterrows():
        age = int(r["age_in_season"])
        traj[age] = (
            float(r["highest_level"]) / 13.0,
            min(float(r["total_minutes"]), 2000) / 2000,
            float(r["avg_score"]) / 0.7 if pd.notna(r.get("avg_score")) else 0.0,
        )
    if len(traj) >= 3:
        pro_age_traj[pid] = traj

print(f"  {len(pro_age_traj)} pro-graczy z ≥3 sezonami w trajektorii wiekowej")


def match_by_age(cand_ages, cand_vecs):
    """
    cand_ages: lista wieków [age_t-2, age_t-1, age_t] (rosnąco)
    cand_vecs: lista (level/13, min/2000, score/0.7) dla tych wieków

    Zwraca top 3 (pid, dist) różnych pid których trajektorie w TYCH wiekach
    są najbliższe.
    """
    cand_arr = np.array(cand_vecs)  # (3, 3)
    results = []
    for pid, traj in pro_age_traj.items():
        # Czy pro ma dane w TYCH SAMYCH wiekach?
        if not all(a in traj for a in cand_ages):
            continue
        pro_arr = np.array([traj[a] for a in cand_ages])  # (3, 3)
        dist = np.linalg.norm(cand_arr - pro_arr)
        results.append((pid, dist))
    results.sort(key=lambda x: x[1])
    # Top 3 różnych pid (już są różne, sortowanie jest po dist)
    return results[:3]


# ══════════════════════════════════════════════════════════════════════════════
# 6. PĘTLA KANDYDATÓW
# ══════════════════════════════════════════════════════════════════════════════

print("\nPrzetwarzam kandydatów...")
t0 = time.time()
total = cand_seasons["player_id"].nunique()
results = []

for idx, (pid, grp) in enumerate(cand_seasons.groupby("player_id"), 1):
    if idx % 5000 == 0:
        el = time.time() - t0
        print(f"  {idx}/{total} ({idx*100/total:.0f}%) | {el:.0f}s, eta {el/idx*(total-idx):.0f}s")

    grp = grp.sort_values("age_in_season").reset_index(drop=True)
    if len(grp) == 0:
        continue

    name = grp["player_name"].iloc[0] if "player_name" in grp.columns else str(pid)
    cur = grp.iloc[-1]
    cur_age = int(cur["age_in_season"])
    cur_level = cur["highest_level"]
    if pd.isna(cur_level):
        continue
    cur_level = int(cur_level)

    if cur_age > 26 or cur_age < 15:
        continue
    if cur_level >= 12:
        continue
    if int(cur["total_minutes"]) < 90:
        continue

    cur_min = int(cur["total_minutes"])
    cur_games = int(cur["matches_played"])
    cur_score = float(cur["avg_score"]) if pd.notna(cur["avg_score"]) else 0.0
    cur_activity = str(cur["activity_status"])

    # Klub: pokaż "dominant → last" jeśli różne, inaczej tylko klub
    cur_club_main = str(cur["club_name"]) if pd.notna(cur["club_name"]) else "—"
    cur_club_dom = str(cur.get("club_name_dominant", cur_club_main))
    if cur_club_dom and cur_club_dom != cur_club_main and cur_club_dom != "—":
        cur_club_display = f"{cur_club_dom} → {cur_club_main}"
    else:
        cur_club_display = cur_club_main

    # Pattern: ostatnie 3 sezony (wieki)
    last3 = grp.tail(3)
    cand_ages = [int(r["age_in_season"]) for _, r in last3.iterrows()]
    cand_vecs = [
        (
            float(r["highest_level"]) / 13.0,
            min(float(r["total_minutes"]), 2000) / 2000,
            float(r["avg_score"]) / 0.7 if pd.notna(r.get("avg_score")) else 0.0,
        )
        for _, r in last3.iterrows()
    ]
    if len(cand_ages) < 3:
        # Padd najwcześniejszym powtórzonym (rzadkie)
        while len(cand_ages) < 3:
            cand_ages.insert(0, cand_ages[0] - 1)
            cand_vecs.insert(0, cand_vecs[0])

    matches = match_by_age(cand_ages, cand_vecs)

    # Ocena
    lvl_s = (cur_level - 4) / 9.0
    min_s = min(cur_min, 2500) / 2500
    scr_s = max(0, min(cur_score / 0.6, 1.0))
    act_map = {"podstawowy": 1.0, "regularny": 0.7, "rezerwowy": 0.4,
               "sporadyczny": 0.1, "nieznana": 0.5}
    act_s = act_map.get(cur_activity, 0.5)
    match_s = max(0, 1 - matches[0][1] / 1.5) if matches else 0
    raw = 0.25*lvl_s + 0.20*min_s + 0.15*scr_s + 0.15*act_s + 0.25*match_s
    ocena = round(raw * 100, 1)

    if   ocena >= 75: status = "★★★ Wyjątkowy"
    elif ocena >= 60: status = "★★ Wysoki"
    elif ocena >= 45: status = "★ Obiecujący"
    elif ocena >= 30: status = "Obserwacja"
    else:             status = "Poniżej"

    # Historia kandydata
    cand_history = []
    prev_club = None
    for _, r in grp.sort_values("age_in_season").tail(6).iterrows():
        club = str(r["club_name"]) if pd.notna(r["club_name"]) else "—"
        club_dom = str(r.get("club_name_dominant", club))
        # W jednym sezonie może być zmiana: pokaż "A → B"
        if club_dom and club_dom != club and club_dom != "—":
            club_display = f"{club_dom} → {club}"
        else:
            club_display = club
        cand_history.append({
            "age":         int(r["age_in_season"]),
            "level":       int(r["highest_level"]) if pd.notna(r["highest_level"]) else None,
            "leagues":     str(r["leagues_str"]),
            "minutes":     int(r["total_minutes"]),
            "matches":     int(r["matches_played"]),
            "score":       round(float(r["avg_score"]), 3) if pd.notna(r.get("avg_score")) else None,
            "activity":    str(r["activity_status"]),
            "club":        club_display,
            "changed":     prev_club is not None and prev_club != club,
        })
        prev_club = club

    # Historia pro: pełna trajektoria po wieku
    pro_histories = []
    for pid_match, dist in matches:
        if pid_match not in pro_by_player:
            continue
        grp_m = pro_by_player[pid_match].sort_values("age_in_season")
        debut_level = int(grp_m["debut_league_level"].iloc[0])

        pro_hist = []
        prev_club = None
        for _, r in grp_m.iterrows():
            club = str(r["club_name"]) if pd.notna(r["club_name"]) else "—"
            pro_hist.append({
                "age":      int(r["age_in_season"]),
                "level":    int(r["highest_level"]),
                "leagues":  str(r["leagues_str"]),
                "minutes":  int(r["total_minutes"]),
                "matches":  int(r["matches_played"]),
                "score":    round(float(r["avg_score"]), 3) if pd.notna(r.get("avg_score")) else None,
                "activity": str(r["activity_status"]),
                "club":     club,
                "changed":  prev_club is not None and prev_club != club,
                "is_match_window": int(r["age_in_season"]) in cand_ages,
            })
            prev_club = club

        ref_row = top_score[top_score["player_id"] == pid_match]
        pro_histories.append({
            "pid":         pid_match,
            "name":        grp_m["player_name"].iloc[0],
            "debut_age":   int(grp_m["age_at_debut"].iloc[0]),
            "debut_level": lvl_name(debut_level),
            "best_score":  float(ref_row["best_overall_score"].iloc[0]) if len(ref_row) > 0 else None,
            "history":     pro_hist,
            "dist":        round(dist, 3),
        })

    results.append({
        "player_id":   pid,
        "name":        name,
        "age":         cur_age,
        "level":       cur_level,
        "level_name":  lvl_name(cur_level) + (" [J]" if bool(cur["highest_is_junior"]) else ""),
        "club":        cur_club_display,
        "minutes":     cur_min,
        "matches":     cur_games,
        "score":       round(cur_score, 3),
        "activity":    cur_activity,
        "ocena":       ocena,
        "status":      status,
        "history":     cand_history,
        "pro_matches": pro_histories,
        "cand_ages":   cand_ages,  # zapamiętaj dla highlight
    })

results.sort(key=lambda x: -x["ocena"])
print(f"  {len(results)} kandydatów po filtrach ({time.time()-t0:.0f}s)")


# ══════════════════════════════════════════════════════════════════════════════
# 7. CANDIDATE_MATCHES.XLSX
# ══════════════════════════════════════════════════════════════════════════════

print("\nBuduję candidate_matches.xlsx...")
t0 = time.time()


def calc_similarity(cand_hist, pro_hist, cand_ages):
    """
    Liczy % zgodności trajektorii kandydata vs pro w 4 wiekach kandydata.

    Wraca dict: {pct_level, pct_min, pct_score, pct_total} — wszystko 0-100.
    Liczone tylko po wiekach gdzie OBA mają dane (intersection).

    pct = 100 * (1 - |diff_normalized|)
      pct_level: diff / 13  (max różnica = 13 poziomów)
      pct_min:   diff / 2000 (cap minut)
      pct_score: diff / 0.7  (cap score)
    """
    # Map age → wartości dla obu
    cand_by_age = {h["age"]: h for h in cand_hist}
    pro_by_age  = {h["age"]: h for h in pro_hist}

    diffs_level, diffs_min, diffs_score = [], [], []

    for age in cand_ages:
        c = cand_by_age.get(age)
        p = pro_by_age.get(age)
        if c is None or p is None:
            continue
        # Level
        c_lvl = c.get("level")
        p_lvl = p.get("level")
        if c_lvl is not None and p_lvl is not None:
            diffs_level.append(abs(c_lvl - p_lvl) / 13)
        # Minutes
        c_min = c.get("minutes")
        p_min = p.get("minutes")
        if c_min is not None and p_min is not None:
            cm = min(c_min, 2000)
            pm = min(p_min, 2000)
            diffs_min.append(abs(cm - pm) / 2000)
        # Score (może być NaN dla niektórych — wtedy pomijaj)
        c_scr = c.get("score")
        p_scr = p.get("score")
        # 'score' nie ma w cand_hist / pro_hist bezpośrednio — musimy dodać
        # (patrz krok niżej — modyfikuję pętlę żeby zachowała 'score')

    pct_level = round(100 * (1 - sum(diffs_level) / len(diffs_level)), 1) if diffs_level else None
    pct_min   = round(100 * (1 - sum(diffs_min) / len(diffs_min)), 1) if diffs_min else None
    pct_score = None  # wypełnimy po dodaniu score do hist

    return {
        "pct_level": pct_level,
        "pct_min":   pct_min,
        "pct_score": pct_score,
        "n_compared": len(diffs_level),
    }


def calc_similarity_v2(cand_hist, pro_hist, cand_ages):
    """
    Wersja używająca pełnych dict-ów z 'score'.
    """
    cand_by_age = {h["age"]: h for h in cand_hist}
    pro_by_age  = {h["age"]: h for h in pro_hist}

    diffs_level, diffs_min, diffs_score = [], [], []

    for age in cand_ages:
        c = cand_by_age.get(age)
        p = pro_by_age.get(age)
        if c is None or p is None:
            continue
        if c.get("level") is not None and p.get("level") is not None:
            diffs_level.append(abs(c["level"] - p["level"]) / 13)
        if c.get("minutes") is not None and p.get("minutes") is not None:
            cm = min(c["minutes"], 2000)
            pm = min(p["minutes"], 2000)
            diffs_min.append(abs(cm - pm) / 2000)
        if c.get("score") is not None and p.get("score") is not None:
            diffs_score.append(min(abs(c["score"] - p["score"]) / 0.7, 1.0))

    pct_level = round(100 * (1 - sum(diffs_level) / len(diffs_level)), 1) if diffs_level else None
    pct_min   = round(100 * (1 - sum(diffs_min)   / len(diffs_min)),   1) if diffs_min   else None
    pct_score = round(100 * (1 - sum(diffs_score) / len(diffs_score)), 1) if diffs_score else None

    parts = [p for p in [pct_level, pct_min, pct_score] if p is not None]
    pct_total = round(sum(parts) / len(parts), 1) if parts else None

    return {
        "pct_level":  pct_level,
        "pct_min":    pct_min,
        "pct_score":  pct_score,
        "pct_total":  pct_total,
        "n_compared": len(diffs_level),
    }


def build_ranking_sheet(wb, results):
    """
    Ranking + trajektorie: dla każdego kandydata pokazujemy 4 sezony (s, s-1, s-2, s-3)
    KANDYDATA, te same 4 wieki PRO-MATCH'a, i % zgodności.
    """
    ws = wb.active
    ws.title = "Ranking"

    # 4 sezony: s (najnowszy), s-1, s-2, s-3 (najstarszy)
    # W kolumnach: od s-3 do s, czyli chronologicznie od najstarszego
    SEZON_LABELS = ["s-3", "s-2", "s-1", "s"]  # wiek rosnący lewo→prawo

    # ── Nagłówki ──────────────────────────────────────────────────────────────
    headers = [
        "#", "Zawodnik", "Klub", "Wiek", "Obecny poziom",
        "Minuty", "Mecze", "Score", "Aktywność",
        "Ocena", "Status",
        "Match #1 (główny)", "M1 wiek deb.",
    ]
    # 4 sezony kandydata × 4 parametry = 16
    for s in SEZON_LABELS:
        headers += [f"K {s}: Wiek", f"K {s}: Liga", f"K {s}: Min", f"K {s}: Score"]
    # 4 sezony pro × 4 parametry = 16
    for s in SEZON_LABELS:
        headers += [f"M1 {s}: Wiek", f"M1 {s}: Liga", f"M1 {s}: Min", f"M1 {s}: Score"]
    # 4 procenty zgodności
    headers += ["% level", "% min", "% score", "% total"]
    # Backupy (tylko nazwiska + dystans)
    headers += ["Match #2 (backup)", "M2 dystans", "Match #3 (backup)", "M3 dystans"]

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_CENTER

    # ── Wiersze ───────────────────────────────────────────────────────────────
    for i, r in enumerate(results, start=1):
        row_num = i + 1
        row = [
            i, r["name"], r["club"], r["age"], r["level_name"],
            r["minutes"], r["matches"], r["score"], r["activity"],
            r["ocena"], r["status"],
        ]
        # Match #1 info
        main = r["pro_matches"][0] if r["pro_matches"] else None
        if main:
            row += [main["name"], main["debut_age"]]
        else:
            row += ["—", "—"]

        # Ostatnie 4 sezony kandydata (od najstarszego)
        cand_last4 = r["history"][-4:]  # już posortowane chronologicznie
        # Padd jeśli mniej niż 4 — pustkami od początku
        while len(cand_last4) < 4:
            cand_last4.insert(0, None)

        cand_ages_4 = [h["age"] if h else None for h in cand_last4]

        for h in cand_last4:
            if h is None:
                row += ["—", "—", "—", "—"]
            else:
                row += [
                    h["age"],
                    lvl_name(h["level"]) if h.get("level") is not None else "—",
                    h["minutes"],
                    h["score"] if h.get("score") is not None else "—",
                ]

        # 4 sezony pro w TYCH SAMYCH WIEKACH co kandydat
        if main:
            pro_by_age = {h["age"]: h for h in main["history"]}
            for age in cand_ages_4:
                if age is None:
                    row += ["—", "—", "—", "—"]
                else:
                    p = pro_by_age.get(age)
                    if p is None:
                        row += [age, "(brak)", "—", "—"]
                    else:
                        row += [
                            p["age"],
                            lvl_name(p["level"]),
                            p["minutes"],
                            p["score"] if p.get("score") is not None else "—",
                        ]

            # Procenty zgodności
            sim = calc_similarity_v2(r["history"], main["history"],
                                     [a for a in cand_ages_4 if a is not None])
            row += [
                sim["pct_level"] if sim["pct_level"] is not None else "—",
                sim["pct_min"]   if sim["pct_min"]   is not None else "—",
                sim["pct_score"] if sim["pct_score"] is not None else "—",
                sim["pct_total"] if sim["pct_total"] is not None else "—",
            ]
        else:
            row += ["—"] * 16  # 4 sezony × 4 parametry
            row += ["—"] * 4   # 4 procenty

        # Backupy
        for j in [1, 2]:
            if len(r["pro_matches"]) > j:
                m = r["pro_matches"][j]
                row += [m["name"], m["dist"]]
            else:
                row += ["—", "—"]

        # Zapisz wiersz
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=row_num, column=col_idx, value=val)
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT if isinstance(val, str) else ALIGN_CENTER

        # Kolorowanie sekcji
        # Sekcja KANDYDAT (kolumny 14-29, czyli kolumny "K s-3..." to col 14, 16 sekcji)
        # info kandydata + match name = 13 kolumn
        # K-sekcja: 4 sezony × 4 param = 16 kolumn → 14..29
        for col_idx in range(14, 14 + 16):
            ws.cell(row=row_num, column=col_idx).fill = FILL_CAND
        # M1-sekcja: 16 kolumn → 30..45
        for col_idx in range(30, 30 + 16):
            ws.cell(row=row_num, column=col_idx).fill = FILL_MAIN
        # Procenty: 46..49
        for col_idx in range(46, 50):
            ws.cell(row=row_num, column=col_idx).fill = FILL_MAIN
        # Backupy: 50+
        for col_idx in range(50, 54):
            ws.cell(row=row_num, column=col_idx).fill = FILL_BACKUP

    # ── Kolorowanie nagłówków sekcji ──────────────────────────────────────────
    # Drugi rząd nagłówków byłby ładny ale openpyxl jest średni w merge — zostawmy
    # tylko fill kolorystyczny w wierszu 1 nagłówków po kolumnach
    for col_idx in range(14, 14 + 16):
        ws.cell(row=1, column=col_idx).fill = PatternFill("solid", start_color="C5A800")
    for col_idx in range(30, 30 + 16):
        ws.cell(row=1, column=col_idx).fill = PatternFill("solid", start_color="305496")
    for col_idx in range(46, 50):
        ws.cell(row=1, column=col_idx).fill = PatternFill("solid", start_color="A5A5A5")
    for col_idx in range(50, 54):
        ws.cell(row=1, column=col_idx).fill = PatternFill("solid", start_color="808080")

    # ── Szerokości kolumn ─────────────────────────────────────────────────────
    widths = [
        5, 24, 32, 6, 22,    # # nazwisko klub wiek poziom
        8, 7, 8, 14,         # min mecze score aktywność
        8, 16,               # ocena status
        22, 11,              # match #1 + wiek deb.
    ]
    # 4 sezony kandydata × 4 param
    for _ in range(4):
        widths += [6, 22, 8, 8]
    # 4 sezony match × 4 param
    for _ in range(4):
        widths += [6, 22, 8, 8]
    # 4 procenty
    widths += [9, 9, 9, 10]
    # Backupy
    widths += [22, 11, 22, 11]

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"  # zamroź # i nazwisko


def format_cell(h, param):
    if h is None:
        return "—"
    if param == "Wiek":         return h["age"]
    if param == "Liga(i)":      return h["leagues"]
    if param == "Łącznie min":  return h["minutes"]
    if param == "Mecze":        return h["matches"]
    if param == "Aktywność":    return h["activity"]
    if param == "Klub":         return h["club"]
    return "—"


def build_comparison_sheet(wb, results, n_top=100):
    """
    Layout: 6 kolumn sezonowych = 6 KOLEJNYCH wieków zawodnika (od najstarszego).
    Każdy zawodnik (kandydat + 3 pro) ma 6 wierszy parametrów.
    """
    ws = wb.create_sheet("Top 100 — szczegóły")

    PARAMS = ["Wiek", "Liga(i)", "Łącznie min", "Mecze", "Aktywność", "Klub"]
    N_COLS = 6  # sezonów (kolumn) na osobę

    row = 1
    for rank_idx, r in enumerate(results[:n_top], start=1):
        # Tytuł kandydata
        title = (f"#{rank_idx}  {r['name']}  ({r['age']} lat, {r['level_name']})  •  "
                 f"{r['club']}  •  Ocena: {r['ocena']}  {r['status']}")
        c = ws.cell(row=row, column=1, value=title)
        c.font = FONT_TITLE; c.fill = FILL_HEADER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS + 1)
        row += 1

        # Nagłówek wieków kandydata
        cand_history = r["history"]  # max 6 wpisów chronologicznie
        cand_ages_match = r["cand_ages"]  # 3 wieki do highlight
        ws.cell(row=row, column=1, value="").fill = FILL_CAND
        for i, h in enumerate(cand_history[-N_COLS:], start=2):
            label = f"Wiek {h['age']}"
            cc = ws.cell(row=row, column=i, value=label)
            cc.font = FONT_BOLD; cc.alignment = ALIGN_CENTER; cc.fill = FILL_CAND
        row += 1

        # Wiersze parametrów kandydata
        for param_idx, param in enumerate(PARAMS):
            label = "KANDYDAT  " + param if param_idx == 0 else "  " + param
            c = ws.cell(row=row, column=1, value=label)
            c.font = FONT_BOLD if param_idx == 0 else FONT_BODY
            c.fill = FILL_CAND
            for i, h in enumerate(cand_history[-N_COLS:], start=2):
                val = format_cell(h, param)
                cc = ws.cell(row=row, column=i, value=val)
                cc.font = FONT_BODY
                cc.alignment = ALIGN_LEFT if isinstance(val, str) else ALIGN_CENTER
                cc.fill = FILL_CAND
            row += 1

        # Bloki pro-matchów
        for m_idx, match in enumerate(r["pro_matches"][:3]):
            is_main = m_idx == 0
            section = "MAIN MATCH" if is_main else f"BACKUP #{m_idx+1}"
            fill = FILL_MAIN if is_main else FILL_BACKUP

            title_text = (f"{section}:  {match['name']}  "
                          f"(debiut w {match['debut_age']}, {match['debut_level']})  •  "
                          f"best score: {match['best_score']:.3f}  •  "
                          f"dystans: {match['dist']}")
            c = ws.cell(row=row, column=1, value=title_text)
            c.font = FONT_BOLD; c.fill = fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS + 1)
            row += 1

            # Wiersz wieków pro (te same wieki co kandydata jeśli to możliwe)
            # Bierzemy 6 ostatnich wieków pro żeby pokazać pełną trajektorię
            pro_history = match["history"]
            pro_history_show = pro_history[-N_COLS:] if len(pro_history) >= N_COLS else pro_history
            ws.cell(row=row, column=1, value="").fill = fill
            for i, h in enumerate(pro_history_show, start=2):
                label = f"Wiek {h['age']}"
                cc = ws.cell(row=row, column=i, value=label)
                cc.font = FONT_BOLD; cc.alignment = ALIGN_CENTER; cc.fill = fill
            row += 1

            # Wiersze parametrów pro
            for param_idx, param in enumerate(PARAMS):
                label = f"{section}  {param}" if param_idx == 0 else "  " + param
                c = ws.cell(row=row, column=1, value=label)
                c.font = FONT_BOLD if param_idx == 0 else FONT_BODY
                c.fill = fill
                for i, h in enumerate(pro_history_show, start=2):
                    val = format_cell(h, param)
                    cc = ws.cell(row=row, column=i, value=val)
                    cc.font = FONT_MATCH_HIGHLIGHT if h.get("is_match_window") else FONT_BODY
                    cc.alignment = ALIGN_LEFT if isinstance(val, str) else ALIGN_CENTER
                    cc.fill = fill
                row += 1

        row += 1  # separator

    ws.column_dimensions["A"].width = 28
    for col in range(2, N_COLS + 2):
        ws.column_dimensions[get_column_letter(col)].width = 32
    ws.freeze_panes = "B2"


wb_cand = Workbook()
build_ranking_sheet(wb_cand, results)
build_comparison_sheet(wb_cand, results, n_top=100)
wb_cand.save(DATA_DIR / "candidate_matches.xlsx")
print(f"  Zapisano: data/candidate_matches.xlsx ({time.time()-t0:.1f}s)")

print(f"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  GOTOWE                                                                      ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  data/pro_career_paths.xlsx                                                  ║
║    • Pro debiut (≥{MIN_SEASONS_FOR_PRO} sezony):  ...{' '*36}║
║    • Detale:      {len(pro_seasons):>5} wierszy{' '*49}║
║                                                                              ║
║  data/candidate_matches.xlsx                                                 ║
║    • Ranking:     {len(results):>5} kandydatów{' '*45}║
║    • Top 100:     wiek-do-wieku matching (kandydat vs pro w TYCH wiekach)    ║
╚══════════════════════════════════════════════════════════════════════════════╝
""")
